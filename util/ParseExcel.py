# -*- coding: utf-8 -*-

#实现读取excel数据文件代码封装
import openpyxl
from openpyxl.styles import Border,Side,Font
import time

class ParseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color = None)#设置字体颜色
        self.RGBDic = {'red':'FFFF3030','green':'FF008B00'}

    def loadWorkBook(self,excelPathAndName):
        #将文件加载到内存，并获取其workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception,e:
            raise e
        self.excelFile = excelPathAndName

    def getSheetByName(self,sheetName):
        #根据sheet名获取sheet对象
        try:
            #sheet = self.workbook.get_sheet_by_name(sheetName)
            sheet = self.workbook[sheetName]
            return sheet
        except Exception,e:
            raise e

    def getSheetByIndex(self,sheetIndex):
        #根据索引号
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception,e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self,sheet):
        #获取结束行号、有多少行
        return sheet.max_row

    def getColsNumber(self,sheet):
        #获取结束列号，有多少列
        return sheet.max_column

    def getStartRowNumber(self,sheet):
        #获取sheet中数据区域开始的行号
        return sheet.min_row

    def getStartColNumber(self, sheet):
        # 获取sheet中数据区域开始的列号
        return sheet.min_column


    def getRow(self, sheet,rowNo):
        # 获取某行所有数据内容
        try:
            #return sheet.rows[rowNo-1] 返回的是一个生成器类型，所以该方法不行
            return list(sheet.rows)[rowNo - 1]
        except Exception,e:
            raise e


    def getColumn(self,sheet,colNo):
        # 获取某列的所有数据内容
        try:
            #
            #return sheet.columns[colNo-1] #返回的是一个生成器类型，所以该方法不行
            return list(sheet.columns)[colNo-1]
        except Exception,e:
            raise e

    def getCellofValue(self,sheet,coordinate = None,rowNo=None,colNo=None):
        if coordinate!=None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception,e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colNo).value
            except Exception,e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")


    def getCellofObject(self, sheet, coordinate=None, rowNo=None, colNo=None):
        #获取某个单元格的对象
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception, e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colNo)
            except Exception, e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")


    def writeCell(self, sheet,content,coordinate=None, rowNo=None, colNo=None,style=None):
        if coordinate != None:
            try:
                sheet.cell(coordinate=coordinate).value=content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font= Font(color=self.RGBDic[style])
                    self.workbook.save(self.excelFile)
            except Exception, e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = content
                if style is not None:
                    sheet.cell(row=rowNo, column=colNo).font = Font(color=self.RGBDic[style])
                    self.workbook.save(self.excelFile)
            except Exception, e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")


    def writeCellCurrenTime(self, sheet,coordinate=None, rowNo=None, colNo=None, style=None):
        #写入当前的时间
        now = int(time.time())
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate != None:
            try:
                sheet.cell(coordinate=coordinate).value = currentTime
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDic[style])
                    self.workbook.save(self.excelFile)
            except Exception, e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = currentTime
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDic[style])
                    self.workbook.save(self.excelFile)
            except Exception, e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

if __name__ =="__main__":
    pe = ParseExcel()
    #pe.loadWorkBook(u"D:\\PythonWork\\KeyWordsFrameWork\\testData\\126邮箱发送邮件.xlsx")
    pe.loadWorkBook(u"..\\testData\\126邮箱发送邮件.xlsx")
    print pe.getSheetByName(u"测试用例").title
    sheet = pe.getSheetByIndex(0)
    print type(sheet)
    print pe.getColsNumber(sheet)

